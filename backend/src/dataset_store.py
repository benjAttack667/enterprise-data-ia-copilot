"""Chargement sécurisé et gestion du dataset actif."""

from __future__ import annotations

import csv
import os
import re
import threading
import uuid
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd


class DatasetError(ValueError):
    """Erreur de validation ou de lecture d'un dataset."""


@dataclass(frozen=True)
class DatasetLimits:
    """Resource limits applied before a dataframe becomes active."""

    max_rows: int = 100_000
    max_columns: int = 200
    max_cells: int = 2_000_000
    max_xlsx_uncompressed_bytes: int = 50 * 1024 * 1024
    max_xlsx_compression_ratio: float = 100.0
    max_xlsx_entries: int = 1_000


@dataclass(frozen=True)
class DatasetSnapshot:
    """Copie cohérente du dataset et de ses métadonnées."""

    id: str
    name: str
    source: str
    updated_at: str
    context: str
    dataframe: pd.DataFrame

    def metadata(self) -> dict[str, object]:
        """Expose les métadonnées attendues par le frontend."""

        rows, columns = self.dataframe.shape
        return {
            "id": self.id,
            "name": self.name,
            "rows": int(rows),
            "columns": int(columns),
            "source": self.source,
            "updated_at": self.updated_at,
            "context": self.context,
        }


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _display_name(filename: str) -> str:
    stem = Path(filename).stem
    words = re.sub(r"[_-]+", " ", stem).strip()
    return words.title() or "Dataset importé"


def _infer_context(filename: str, columns: list[str]) -> str:
    """Déduit un contexte lisible sans prétendre connaître le métier du client."""

    tokens = " ".join([filename, *columns]).lower()
    if any(token in tokens for token in ("lead", "campaign", "conversion", "revenue")):
        return "Marketing & Sales"
    if any(token in tokens for token in ("supplier", "compliance", "recycl", "packaging")):
        return "Supply Chain & ESG"
    if any(token in tokens for token in ("project", "innovation", "progress", "roi")):
        return "Innovation & PMO"
    if any(token in tokens for token in ("forecast", "margin", "actual", "finance")):
        return "Finance & Performance"
    return "Analyse de données"


def _validate_headers(headers: list[object]) -> list[str]:
    """Reject empty/duplicate headers before Pandas can silently rename them."""

    normalized = ["" if value is None else str(value).strip() for value in headers]
    if not normalized or any(not value for value in normalized):
        raise DatasetError("Chaque colonne doit avoir un nom non vide.")
    if len(normalized) != len(set(normalized)):
        raise DatasetError("Le fichier contient des noms de colonnes dupliqués.")
    return normalized


def _csv_headers(path: Path) -> list[str]:
    """Read only the first CSV record, with the encodings already supported."""

    for encoding in ("utf-8-sig", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as stream:
                return _validate_headers(list(next(csv.reader(stream), [])))
        except UnicodeDecodeError:
            continue
    raise DatasetError("Le fichier CSV utilise un encodage non pris en charge.")


def _validate_xlsx_archive(path: Path, limits: DatasetLimits) -> None:
    """Reject malformed, encrypted or disproportionately expanded XLSX files."""

    if not zipfile.is_zipfile(path):
        raise DatasetError("Le fichier XLSX n'est pas une archive Excel valide.")
    archive_size = max(path.stat().st_size, 1)
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.infolist()
            if len(members) > limits.max_xlsx_entries:
                raise DatasetError(
                    "Le fichier XLSX contient trop d'éléments internes."
                )
            if any(member.flag_bits & 0x1 for member in members):
                raise DatasetError("Les fichiers XLSX chiffrés ne sont pas acceptés.")
            uncompressed_size = sum(member.file_size for member in members)
    except DatasetError:
        raise
    except OSError:
        raise
    except zipfile.BadZipFile as exc:
        raise DatasetError("Le fichier XLSX n'est pas une archive valide.") from exc

    if uncompressed_size > limits.max_xlsx_uncompressed_bytes:
        raise DatasetError("Le contenu XLSX décompressé dépasse la limite configurée.")
    if uncompressed_size / archive_size > limits.max_xlsx_compression_ratio:
        raise DatasetError("Le taux de compression du fichier XLSX est trop élevé.")


def _xlsx_shape_and_headers(
    path: Path, limits: DatasetLimits
) -> tuple[int, int, list[str]]:
    """Inspect worksheet dimensions in read-only mode before Pandas parsing."""

    from openpyxl import load_workbook

    _validate_xlsx_archive(path, limits)
    try:
        with path.open("rb") as binary_stream:
            workbook = load_workbook(
                binary_stream,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
            try:
                worksheet = workbook.active
                row_count = max(int(worksheet.max_row or 0) - 1, 0)
                column_count = int(worksheet.max_column or 0)
                headers = _validate_headers(
                    list(
                        next(
                            worksheet.iter_rows(
                                min_row=1,
                                max_row=1,
                                values_only=True,
                            ),
                            (),
                        )
                    )
                )
            finally:
                workbook.close()
    except DatasetError:
        raise
    except OSError:
        raise
    except Exception as exc:
        raise DatasetError("Le fichier XLSX ne peut pas être inspecté.") from exc
    return row_count, column_count, headers


def _validate_shape(rows: int, columns: int, limits: DatasetLimits) -> None:
    """Enforce bounded dataframe dimensions with stable public errors."""

    if columns > limits.max_columns:
        raise DatasetError(
            f"Le dataset dépasse la limite de {limits.max_columns} colonnes."
        )
    if rows > limits.max_rows:
        raise DatasetError(
            f"Le dataset dépasse la limite de {limits.max_rows} lignes."
        )
    if rows * columns > limits.max_cells:
        raise DatasetError(
            f"Le dataset dépasse la limite de {limits.max_cells} cellules."
        )


def _read_dataframe_path(
    filename: str, path: Path, limits: DatasetLimits
) -> pd.DataFrame:
    """Contrôle les en-têtes avant que Pandas ne renomme les doublons.

    ``read_csv`` et ``read_excel`` rendent par défaut les noms dupliqués
    artificiellement uniques (par exemple ``name.1``). Le contrôle doit donc
    lire la première ligne directement depuis le format source.
    """

    suffix = Path(filename).suffix.lower()
    try:
        if suffix == ".csv":
            headers = _csv_headers(path)
            _validate_shape(0, len(headers), limits)
            max_rows_by_cells = limits.max_cells // len(headers)
            read_limit = min(limits.max_rows, max_rows_by_cells)
            try:
                dataframe = pd.read_csv(path, nrows=read_limit + 1)
            except UnicodeDecodeError:
                dataframe = pd.read_csv(
                    path,
                    encoding="latin-1",
                    nrows=read_limit + 1,
                )
        elif suffix == ".xlsx":
            rows, columns, _headers = _xlsx_shape_and_headers(path, limits)
            _validate_shape(rows, columns, limits)
            with path.open("rb") as binary_stream:
                dataframe = pd.read_excel(
                    binary_stream,
                    engine="openpyxl",
                    nrows=limits.max_rows + 1,
                )
        else:
            raise DatasetError("Format non pris en charge. Utilisez un fichier CSV ou XLSX.")
    except DatasetError:
        raise
    except OSError:
        raise
    except Exception as exc:  # Pandas remonte plusieurs familles d'erreurs de parse.
        raise DatasetError(f"Le fichier ne peut pas être lu : {exc}") from exc

    if dataframe.empty or dataframe.shape[1] == 0:
        raise DatasetError("Le fichier ne contient aucune donnée exploitable.")
    _validate_shape(int(dataframe.shape[0]), int(dataframe.shape[1]), limits)
    dataframe.columns = [str(column).strip() for column in dataframe.columns]
    if any(not column for column in dataframe.columns):
        raise DatasetError("Chaque colonne doit avoir un nom non vide.")
    if dataframe.columns.duplicated().any():
        raise DatasetError("Le fichier contient des noms de colonnes dupliqués.")
    return dataframe


class DatasetStore:
    """Conserve un dataset actif et renvoie toujours des copies en lecture."""

    def __init__(
        self,
        samples_dir: Path,
        uploads_dir: Path,
        limits: DatasetLimits | None = None,
    ) -> None:
        self.samples_dir = samples_dir
        self.uploads_dir = uploads_dir
        self.limits = limits or DatasetLimits()
        self._lock = threading.RLock()
        self.uploads_dir.mkdir(parents=True, exist_ok=True)
        self._cleanup_staged_uploads()
        self._prune_final_uploads()
        self._active = self._load_default()

    def _load_default(self) -> DatasetSnapshot:
        default_path = self.samples_dir / "marketing_leads.csv"
        if not default_path.is_file():
            raise RuntimeError(f"Dataset de démonstration absent : {default_path}")
        dataframe = _read_dataframe_path(default_path.name, default_path, self.limits)
        updated_at = datetime.fromtimestamp(
            default_path.stat().st_mtime, tz=timezone.utc
        ).isoformat()
        return DatasetSnapshot(
            id="marketing-leads",
            name="Marketing Leads",
            source="sample",
            updated_at=updated_at,
            context="Marketing & Sales",
            dataframe=dataframe,
        )

    def _final_upload_paths(self) -> list[Path]:
        return [
            path
            for pattern in ("*.csv", "*.xlsx")
            for path in self.uploads_dir.glob(pattern)
            if path.is_file()
        ]

    def _cleanup_staged_uploads(self) -> None:
        """Remove interrupted uploads left by a stopped process."""

        for path in self.uploads_dir.glob(".upload-*.part"):
            path.unlink(missing_ok=True)

    def _prune_final_uploads(self, keep: Path | None = None) -> None:
        """Keep exactly one final upload, preferring the newly active file."""

        paths = self._final_upload_paths()
        if keep is None and paths:
            keep = max(paths, key=lambda path: path.stat().st_mtime_ns)
        for path in paths:
            if keep is not None and path == keep:
                continue
            path.unlink(missing_ok=True)

    @staticmethod
    def _copy_snapshot(snapshot: DatasetSnapshot) -> DatasetSnapshot:
        return DatasetSnapshot(
            id=snapshot.id,
            name=snapshot.name,
            source=snapshot.source,
            updated_at=snapshot.updated_at,
            context=snapshot.context,
            dataframe=snapshot.dataframe.copy(deep=True),
        )

    def get_active(self) -> DatasetSnapshot:
        """Retourne une copie, empêchant un endpoint de modifier l'état partagé."""

        with self._lock:
            return self._copy_snapshot(self._active)

    def create_staged_upload(self) -> Path:
        """Reserve a private path used by the endpoint for incremental writes."""

        path = self.uploads_dir / f".upload-{uuid.uuid4().hex}.part"
        descriptor = os.open(path, os.O_WRONLY | os.O_CREAT | os.O_EXCL, 0o600)
        os.close(descriptor)
        return path

    def discard_staged_upload(self, staged_path: Path) -> None:
        """Remove only a staging file contained in the configured upload root."""

        resolved = staged_path.resolve()
        root = self.uploads_dir.resolve()
        if root not in resolved.parents or not resolved.name.startswith(".upload-"):
            return
        try:
            resolved.unlink(missing_ok=True)
        except OSError:
            pass

    def activate_staged_upload(
        self, filename: str, staged_path: Path
    ) -> DatasetSnapshot:
        """Validate a staged file, atomically retain it and activate its data.

        La sauvegarde n'a lieu qu'après un parsing réussi, ce qui évite de garder
        des fichiers invalides dans ``data/uploads``.
        """

        safe_original_name = Path(filename).name
        suffix = Path(safe_original_name).suffix.lower()
        if suffix not in {".csv", ".xlsx"}:
            raise DatasetError("Format non pris en charge. Utilisez un fichier CSV ou XLSX.")
        staged_path = staged_path.resolve()
        uploads_root = self.uploads_dir.resolve()
        if uploads_root not in staged_path.parents:
            raise DatasetError("Chemin temporaire invalide.")
        dataframe = _read_dataframe_path(safe_original_name, staged_path, self.limits)
        dataset_id = uuid.uuid4().hex
        stored_name = f"{dataset_id}{suffix}"
        stored_path = (self.uploads_dir / stored_name).resolve()
        if uploads_root not in stored_path.parents:
            raise DatasetError("Nom de fichier invalide.")

        snapshot = DatasetSnapshot(
            id=dataset_id,
            name=_display_name(safe_original_name),
            source="upload",
            updated_at=_utc_now(),
            context=_infer_context(safe_original_name, list(dataframe.columns)),
            dataframe=dataframe,
        )
        with self._lock:
            os.replace(staged_path, stored_path)
            try:
                self._prune_final_uploads(keep=stored_path)
            except OSError as prune_error:
                try:
                    stored_path.unlink(missing_ok=True)
                except OSError as rollback_error:
                    raise rollback_error from prune_error
                raise
            # Publish the in-memory snapshot only after disk retention succeeded.
            self._active = snapshot
            return self._copy_snapshot(snapshot)

    def activate_upload(self, filename: str, content: bytes) -> DatasetSnapshot:
        """Compatibility wrapper for trusted in-process callers using bytes."""

        staged_path = self.create_staged_upload()
        try:
            staged_path.write_bytes(content)
            return self.activate_staged_upload(filename, staged_path)
        finally:
            self.discard_staged_upload(staged_path)

    def storage_metrics(self) -> dict[str, int]:
        """Return bounded aggregate usage without exposing names or paths."""

        with self._lock:
            files = self._final_upload_paths()
            return {
                "files": len(files),
                "bytes": sum(path.stat().st_size for path in files),
                "max_files": 1,
            }
